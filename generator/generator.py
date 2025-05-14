import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

def analyze_instance(file_path):
    """Analyze the book scanning instance file and return metrics"""
    try:
        with open(file_path, 'r') as f:
            first_line = f.readline().strip().split()
            B = int(first_line[0])
            L = int(first_line[1])
            D = int(first_line[2])

            book_scores = list(map(int, f.readline().strip().split()))

            total_books_in_libraries = 0
            total_signup_time = 0
            total_shipping_capacity = 0

            for _ in range(L):
                lib_line = f.readline().strip().split()
                N_j = int(lib_line[0])
                T_j = int(lib_line[1])
                M_j = int(lib_line[2])

                _ = f.readline()

                total_books_in_libraries += N_j
                total_signup_time += T_j
                total_shipping_capacity += M_j

        avg_book_score = sum(book_scores) / B if B > 0 else 0
        avg_books_per_library = total_books_in_libraries / L if L > 0 else 0
        avg_signup_time = total_signup_time / L if L > 0 else 0
        avg_shipping_capacity = total_shipping_capacity / L if L > 0 else 0

        return {
            "No. of books": B,
            "No. of libraries": L,
            "No. of days": D,
            "Average book score": round(avg_book_score, 2),
            "Average no. of books per library": round(avg_books_per_library, 2),
            "Average sign up time": round(avg_signup_time, 2),
            "Average No. of Books Shipped per Library per Day": round(avg_shipping_capacity, 2)
        }

    except FileNotFoundError:
        print(f"Error: Input file not found at {file_path}")
        return None
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        return None

def create_excel_report(input_file_paths, output_file_path):
    """Create Excel report with metrics for all instances"""
    try:
        # If output file exists, load it, otherwise create a new DataFrame
        if os.path.exists(output_file_path):
            df = pd.read_excel(output_file_path)
        else:
            data = {
                "Instance": [],
                "No. of books": [],
                "No. of libraries": [],
                "No. of days": [],
                "Average book score": [],
                "Average no. of books per library": [],
                "Average sign up time": [],
                "Average No. of Books Shipped per Library per Day": []
            }
            df = pd.DataFrame(data)

        # Iterate over all input files
        for input_file in input_file_paths:
            instance_name = os.path.basename(input_file).split('.')[0]
            metrics = analyze_instance(input_file)

            if not metrics:
                continue

            # Find or add a new row for the instance
            if instance_name in df["Instance"].values:
                instance_index = df.index[df["Instance"] == instance_name].tolist()[0]
            else:
                instance_index = len(df)
                df.loc[instance_index, "Instance"] = instance_name

            # Add metrics to the row
            for key, value in metrics.items():
                df.loc[instance_index, key] = value

        # Save DataFrame to Excel
        df.to_excel(output_file_path, index=False)

        # Formatting the Excel file
        wb = load_workbook(output_file_path)
        ws = wb.active

        # Style headers
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_file_path)
        print(f"Successfully generated report: {output_file_path}")
        return True

    except Exception as e:
        print(f"Error generating Excel file: {str(e)}")
        return False

if __name__ == "__main__":
    current_dir = os.path.dirname(os.path.abspath(__file__))
    input_folder = os.path.join(os.path.dirname(current_dir), "input")
    output_folder = current_dir

    # List of instance files to process
    instance_files = [
        # "B5000_L90_D21.txt", "B50000_L400_D28.txt", "B90000_L850_D21.txt", "B95000_L2000_D28.txt", "B100000_L600_D28.txt",
        "a_example.txt", "b_read_on.txt", "c_incunabula.txt", "d_tough_choices.txt", "e_so_many_books.txt", "f_libraries_of_the_world.txt", "switch_books_instance.txt", "UPFIEK.txt"
    ]


    input_file_paths = [os.path.join(input_folder, file) for file in instance_files]
    output_file = os.path.join(output_folder, "library_metrics.xlsx")

    if not all(os.path.exists(file) for file in input_file_paths):
        print("Error: Some input files are missing. Please ensure all instance files exist in the 'input' folder.")
    else:
        success = create_excel_report(input_file_paths=input_file_paths, output_file_path=output_file)

        if success:
            print("Report generated successfully!")
            print(f"Location: {output_file}")
        else:
            print("Failed to generate report")
